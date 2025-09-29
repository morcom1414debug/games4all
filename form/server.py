# server.py - Flask app to receive form submissions and append to data.xlsx
# Place this server.py and your data.xlsx in the same folder, then run:
# pip install flask openpyxl
# python server.py
# Open http://127.0.0.1:5000 in your browser.
from flask import Flask, request, jsonify, send_from_directory
from openpyxl import load_workbook, Workbook
import os

APP = Flask(__name__, static_folder='.', static_url_path='')

DATA_FILE = 'data.xlsx'  # make sure data.xlsx is in the same directory

def ensure_file():
    if not os.path.exists(DATA_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws['A1'] = 'ลำดับที่'
        ws['B1'] = 'ชื่อ-สกุล'
        ws['C1'] = 'เพศ'
        ws['D1'] = 'ความสนใจ'
        wb.save(DATA_FILE)

def get_next_sequence(ws):
    # find last numeric sequence in column A starting from row 2
    last = 0
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        val = row[0].value
        if isinstance(val, (int, float)):
            try:
                n = int(val)
                if n > last:
                    last = n
            except:
                pass
        else:
            # if string that can be int
            try:
                n = int(str(val))
                if n > last:
                    last = n
            except:
                pass
    return last + 1

@APP.route('/')
def index():
    # serve form.html
    return send_from_directory('.', 'form.html')

@APP.route('/submit', methods=['POST'])
def submit():
    data = request.get_json()
    if not data:
        return jsonify(error='ข้อมูลไม่ถูกต้อง'), 400
    fullname = (data.get('fullname') or '').strip()
    gender = (data.get('gender') or '').strip()
    interests = data.get('interests') or []
    if not fullname or not gender or not interests:
        return jsonify(error='กรุณากรอกข้อมูลให้ครบทุกช่อง'), 400

    ensure_file()
    wb = load_workbook(DATA_FILE)
    ws = wb.active

    seq = get_next_sequence(ws)
    next_row = ws.max_row + 1
    # but ensure we write to the correct next empty row (in case there are trailing empties)
    # find first empty row after header
    r = 2
    while ws.cell(row=r, column=1).value not in (None, ''):
        r += 1
    next_row = r

    ws.cell(row=next_row, column=1, value=seq)
    ws.cell(row=next_row, column=2, value=fullname)
    ws.cell(row=next_row, column=3, value=gender)
    ws.cell(row=next_row, column=4, value=','.join(interests))

    wb.save(DATA_FILE)
    return jsonify(success=True, sequence=seq), 200

if __name__ == '__main__':
    ensure_file()
    print('Starting server on http://127.0.0.1:5000')
    APP.run(debug=True)
